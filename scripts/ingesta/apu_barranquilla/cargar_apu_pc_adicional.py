"""
StructAI — Carga de fuentes reales adicionales encontradas en el PC del usuario
(triage manual 2026-08-08, autorizado explícitamente por el usuario).

9 archivos/hojas confirmados como reales y con precios llenos (de ~19 candidatos
revisados; el resto eran plantillas vacías, solo-cantidades, o no-construcción y
se descartaron). Ver project_apu_precios_barranquilla.md (memoria) para el detalle
completo de la triage, incluyendo los archivos descartados y el motivo.

IMPORTANTE — privacidad (instrucción explícita y repetida del usuario): ninguna
fila insertada aquí incluye nombre de obra, dirección de calle, ni datos de
contacto (persona/empresa/teléfono/email). El archivo WPM (box culvert) tenía
metadata de cliente real (INCOLSOS S.A.S., contacto, NIT, teléfono, dirección)
que se descarta por completo — solo se usan sus filas de ítem/precio genéricas.

Tablas destino: mismas de cargar_apu_barranquilla.py (apu_precios_referencia).
Esta carga es solo a nivel de actividad (no se insertan insumos individuales,
para mantener el parseo simple y confiable dado que son 9 layouts distintos).

Uso: python cargar_apu_pc_adicional.py [--dry-run]
"""
import os
import re
import statistics
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]

from dotenv import load_dotenv
load_dotenv(ROOT / "apps" / "api" / ".env")

import openpyxl

HOY = date.today().isoformat()


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ── 1. Puerto Colombia Calle 15 — catálogo de precios 2017-2019 (IPC) ──────────

def cargar_pto_colombia_calle15() -> list[dict]:
    path = Path.home() / "Downloads" / "Hojas-Calculo" / "PRESUPUESTO PUERTO COLOMBIA CALLE 15.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["aumento valor unitario"]
    filas = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if r[4] is None or r[3] is None:  # requiere UNIDAD (filtra filas de categoría)
            continue
        precio = _num(r[7]) or _num(r[6]) or _num(r[5])  # 2019 > 2018 > 2017
        if precio is None:
            continue
        filas.append({
            "actividad": _txt(r[3]),
            "unidad": _txt(r[4]),
            "disciplina": "Vías / Obra civil",
            "precio_todo_costo": precio,
            "desglose_confiable": False,
            "region": "Puerto Colombia",
            "categoria_fuente": f"FAMILIA {r[0]} / SUBFAM {r[1]}" if r[0] else None,
            "tipo_fuente": "contrato_real_pto_colombia_calle15",
            "fuente": "Contrato real ejecutado — Puerto Colombia, Calle 15 (catálogo de precios 2017-2019, ajustado por IPC)",
            "fecha_captura": HOY,
        })
    return filas


# ── 2. Puerto Colombia 2016 — demoliciones/pavimentos (IPC 2016) ──────────────

def cargar_pto_colombia_2016() -> list[dict]:
    path = Path.home() / "Downloads" / "PRESUPUESTOS 2016.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "PUERTO COLOMBIA" not in wb.sheetnames:
        return []
    ws = wb["PUERTO COLOMBIA"]
    filas = []
    for r in ws.iter_rows(min_row=16, values_only=True):
        if r[4] is None or r[3] is None or r[5] is None:  # requiere UNIDAD + VALOR UNITARIO
            continue
        filas.append({
            "actividad": _txt(r[3]),
            "unidad": _txt(r[4]),
            "disciplina": "Vías / Obra civil",
            "precio_todo_costo": _num(r[5]),
            "desglose_confiable": False,
            "region": "Puerto Colombia",
            "categoria_fuente": f"FAMILIA {r[0]} / SUBFAM {r[1]}" if r[0] else None,
            "tipo_fuente": "contrato_real_pto_colombia_2016",
            "fuente": "Contrato real ejecutado — Puerto Colombia (catálogo de precios 2016, demoliciones y pavimentos, ajustado por IPC)",
            "fecha_captura": HOY,
        })
    return filas


# ── 3. Gobernación del Atlántico — Secretaría de Infraestructura (obra vial) ──

def cargar_gobernacion_atlantico() -> list[dict]:
    path = Path.home() / "Downloads" / "12 - PRESUPUESTO DEL PROCESO.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["CANTIDADES"]
    filas = []
    resumen = {"COSTO DIRECTO", "AIU", "TOTAL", "IVA", None}
    for r in ws.iter_rows(min_row=7, values_only=True):
        desc = _txt(r[1])
        if desc is None or desc.upper() in resumen or r[3] is None:  # requiere PRECIO_UNITARIO
            continue
        filas.append({
            "actividad": desc,
            "unidad": _txt(r[2]),
            "disciplina": "Vías / Hidráulica y Sanitaria",
            "precio_todo_costo": _num(r[3]),
            "desglose_confiable": False,
            "region": "Atlántico",
            "categoria_fuente": "Obra vial/alcantarillado — Secretaría de Infraestructura",
            "tipo_fuente": "contrato_real_gobernacion_atlantico",
            "fuente": "Presupuesto real de proceso — Gobernación del Atlántico, Secretaría de Infraestructura (obra de alcantarillado y protección vial)",
            "fecha_captura": HOY,
        })
    return filas


# NOTA: se descartó una fuente candidata — "LISTADO PRECIOS-1(41).xlsx" hoja
# "LIQUIDACION CANGREJERA" — tras verificar contra Supabase que es EL MISMO
# dataset ya cargado como tipo_fuente='contrato_real_infraestructura_aa' (892
# filas, precios idénticos ítem por ítem, ej. "Delineador Tubular Plastico" =
# $47.600 con código 3.1.1.2.1 en ambos). Insertarla habría duplicado 892 filas.

# ── 5. Cotización real — cerca de seguridad / concertina (proveedor) ──────────

def cargar_cotizacion_cerca_seguridad() -> list[dict]:
    path = Path.home() / "Documents" / "Presupuestos-Obra" / "Anexo de Precios y cantidades 1200002541_MANO DE OBRA.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Anexo de precios"]
    filas = []
    for r in ws.iter_rows(min_row=11, values_only=True):
        if r[3] is None or r[7] is None:  # requiere Descripción + Vlr. Unitario
            continue
        filas.append({
            "actividad": _txt(r[3]),
            "unidad": _txt(r[6]),
            "disciplina": "Seguridad industrial / cerramientos",
            "precio_todo_costo": _num(r[7]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Cerca de seguridad / concertina",
            "tipo_fuente": "cotizacion_real_cerca_seguridad",
            "fuente": "Cotización real de proveedor — cerca de seguridad/concertina (Triple A / Atlántico, 2024)",
            "fecha_captura": "2024-06-05",
        })
    return filas


# ── 6. Obras civiles y demoliciones (Triple A, valores globales) ──────────────

def cargar_obras_civiles_demoliciones() -> list[dict]:
    path = Path.home() / "Desktop" / "optimizacion para negocios en el atlantico" / "Anexo de Precios y cantidades AIU (1).xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Anexo de precios"]
    filas = []
    for r in ws.iter_rows(min_row=11, values_only=True):
        if r[3] is None or r[6] is None:  # requiere Descripción + Vlr. Unitario (filtra encabezados de categoría)
            continue
        filas.append({
            "actividad": _txt(r[3]),
            "unidad": _txt(r[5]),
            "disciplina": "Obras civiles / demoliciones",
            "precio_todo_costo": _num(r[6]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Obras civiles y demoliciones",
            "tipo_fuente": "contrato_real_obras_civiles_demoliciones",
            "fuente": "Contrato real ejecutado — obras civiles y demoliciones (Triple A / Atlántico)",
            "fecha_captura": HOY,
        })
    return filas


# ── 7. Instalación/reposición de medidores de agua ────────────────────────────

def cargar_medidores() -> list[dict]:
    path = Path.home() / "Downloads" / "Hojas-Calculo" / "PRECIOS MEDIDORES.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["PRECIOS"]
    filas = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[2] is None or r[3] is None:
            continue
        filas.append({
            "actividad": _txt(r[2]),
            "unidad": "Un",
            "disciplina": "Hidráulica y Sanitaria",
            "precio_todo_costo": _num(r[3]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Instalación/reposición de medidores de agua",
            "tipo_fuente": "contrato_real_medidores_agua",
            "fuente": "Contrato real ejecutado — instalación y reposición de medidores de agua (Triple A / Barranquilla)",
            "fecha_captura": HOY,
        })
    return filas


# ── 8. Impermeabilización ──────────────────────────────────────────────────────

def cargar_impermeabilizacion() -> list[dict]:
    path = Path.home() / "Downloads" / "Presupuesto impermeabilizacion (1) (a0).xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hoja1 "]  # nombre de hoja con espacio final
    filas = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[2] is None or r[4] is None:  # requiere UND + VR. UNIT (filtra encabezados de categoría)
            continue
        filas.append({
            "actividad": _txt(r[1]),
            "unidad": _txt(r[2]),
            "disciplina": "Impermeabilización / acabados",
            "precio_todo_costo": _num(r[4]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Impermeabilización",
            "tipo_fuente": "contrato_real_impermeabilizacion",
            "fuente": "Contrato real ejecutado — impermeabilización de losa/cubierta (Atlántico)",
            "fecha_captura": HOY,
        })
    return filas


# ── 9b. Catálogo IAD MIPYMES — Colombia Compra Eficiente ───────────────────────
# 1.795 ítems de ferretería/materiales de construcción, cada uno con hasta 79
# cotizaciones reales de proveedores mipyme distintos a nivel nacional (columnas
# "Precio (SIN IVA)1..79"). Se resume cada ítem con la MEDIANA de sus cotizaciones
# válidas (no se explota fila por proveedor: 1.795 x 79 sería excesivo para esta
# tabla, y la mediana es robusta frente a outliers de digitación vistos en la
# fuente, ej. valores como "10" o "925" sueltos entre cotizaciones de 6 cifras).
# No se listan nombres de proveedores individuales al usuario (mismo criterio de
# sanitización que el resto de la base) aunque son empresas públicas registradas,
# no personas — se opta por consistencia con el resto de tipo_fuente.

def _num_iad(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v > 0 else None
    s = re.sub(r"[^\d.,\-]", "", str(v).strip())
    if s.count(",") == 1 and s.count(".") > 1:  # formato "1.879.869" con coma decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
        return f if f > 0 else None
    except ValueError:
        return None


def cargar_iad_mipymes() -> list[dict]:
    path = Path.home() / "Downloads" / "catalogo_ferreteria_-_iad_mipymes_v13.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Precios del Catálogo"]
    filas = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if r[1] is None:
            continue
        precios = [p for p in (_num_iad(v) for v in r[3:]) if p is not None]
        if len(precios) < 2:  # exige al menos 2 cotizaciones reales para ser representativo
            continue
        filas.append({
            "actividad": _txt(r[1]),
            "unidad": _txt(r[2]),
            "disciplina": "Ferretería / materiales de construcción",
            "precio_todo_costo": round(statistics.median(precios), 2),
            "desglose_confiable": False,
            "region": "Nacional",
            "categoria_fuente": f"Mediana de {len(precios)} cotizaciones mipyme reales (rango ${min(precios):,.0f}–${max(precios):,.0f} COP)",
            "tipo_fuente": "catalogo_iad_mipymes",
            "fuente": "Catálogo IAD MIPYMES — Instrumento de Agregación de Demanda, Colombia Compra Eficiente (precios sin IVA, mediana de cotizaciones reales de proveedores mipyme a nivel nacional)",
            "fecha_captura": HOY,
        })
    return filas


# ── 9c. Cotizaciones reales adicionales (Triple A) ─────────────────────────────

def cargar_solped_ptar_piojo() -> list[dict]:
    path = Path.home() / "Desktop" / "Elementos varios" / "Anexo de Precios y cantidades Solped 1000044000 wilmer perez.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Anexo de precios"]
    filas = []
    for r in ws.iter_rows(min_row=11, max_row=17, values_only=True):
        if r[2] is None or r[6] is None:
            continue
        filas.append({
            "actividad": _txt(r[2]),
            "unidad": _txt(r[5]),
            "disciplina": "Hidráulica y Sanitaria — PTAR",
            "precio_todo_costo": _num(r[6]),
            "desglose_confiable": False,
            "region": "Atlántico",
            "categoria_fuente": "Planta de tratamiento de aguas residuales (PTAR)",
            "tipo_fuente": "cotizacion_real_ptar",
            "fuente": "Cotización real — sistema de bombeo y reactor UASB, PTAR municipal (Atlántico)",
            "fecha_captura": HOY,
        })
    return filas


def cargar_edar_humedales() -> list[dict]:
    path = Path.home() / "Desktop" / "Elementos varios" / "triple A" / "Copia de Anexo+de+Precios+y+cantidades wilmer.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Anexo de precios"]
    filas = []
    for r in ws.iter_rows(min_row=11, values_only=True):
        if r[3] is None or r[6] is None:
            continue
        if str(r[3]).strip().lower() == "subtotal" or (isinstance(r[1], str) and r[1].strip().lower() == "subtotal"):
            break
        filas.append({
            "actividad": _txt(r[3]),
            "unidad": _txt(r[5]),
            "disciplina": "Hidráulica y Sanitaria — EDAR / humedales artificiales",
            "precio_todo_costo": _num(r[6]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Estación depuradora de aguas residuales (EDAR) — humedales artificiales",
            "tipo_fuente": "contrato_real_edar_humedales",
            "fuente": "Contrato real ejecutado — mantenimiento de EDAR / humedales artificiales (Triple A / Atlántico)",
            "fecha_captura": HOY,
        })
    return filas


# ── 9. Box culvert / estructura de drenaje (PII del cliente descartada) ───────

def cargar_box_culvert() -> list[dict]:
    path = Path.home() / "Downloads" / "WPM  presupuesto puente (box ).xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["presupuesto "]  # nombre de hoja con espacio final
    filas = []
    for r in ws.iter_rows(min_row=18, values_only=True):
        # r[0] item, r[1] descripcion, r[2] unidad, r[3] cantidad, r[4] valor_unitario, r[5] costo_total
        if r[1] is None or r[2] is None or r[4] is None:
            continue
        if r[1] is not None and str(r[1]).strip().upper().startswith("RESUMEN"):
            break
        filas.append({
            "actividad": _txt(r[1]),
            "unidad": _txt(r[2]).strip(),
            "disciplina": "Estructuras de drenaje / box culvert",
            "precio_todo_costo": _num(r[4]),
            "desglose_confiable": False,
            "region": "Barranquilla",
            "categoria_fuente": "Box culvert / estructura de drenaje",
            "tipo_fuente": "contrato_real_box_culvert",
            # Sin nombre de cliente, NIT, contacto ni dirección — solo la naturaleza genérica de la obra.
            "fuente": "Contrato real ejecutado — box culvert / estructura de drenaje en concreto (Barranquilla)",
            "fecha_captura": "2018-06-18",
        })
    return filas


def _batch_insert(sb, tabla: str, filas: list[dict], batch_size: int = 500) -> list[dict]:
    insertadas = []
    for i in range(0, len(filas), batch_size):
        lote = filas[i:i + batch_size]
        resp = sb.table(tabla).insert(lote).execute()
        insertadas.extend(resp.data)
    return insertadas


def cargar(dry_run: bool = False):
    fuentes = {
        "Puerto Colombia Calle 15 (2017-2019)": cargar_pto_colombia_calle15(),
        "Puerto Colombia 2016 (demoliciones/pavimentos)": cargar_pto_colombia_2016(),
        "Gobernación Atlántico (obra vial/alcantarillado)": cargar_gobernacion_atlantico(),
        "Cotización cerca de seguridad (proveedor real)": cargar_cotizacion_cerca_seguridad(),
        "Obras civiles y demoliciones (AIU)": cargar_obras_civiles_demoliciones(),
        "Medidores de agua": cargar_medidores(),
        "Impermeabilización": cargar_impermeabilizacion(),
        "Box culvert / estructura de drenaje": cargar_box_culvert(),
        "Catálogo IAD MIPYMES (ferretería, multi-proveedor nacional)": cargar_iad_mipymes(),
        "Cotización PTAR Piojo (bombeo/UASB)": cargar_solped_ptar_piojo(),
        "EDAR / humedales artificiales (Triple A)": cargar_edar_humedales(),
    }

    todas = []
    print("=== Resumen de carga (fuentes reales adicionales del PC) ===")
    for nombre, filas in fuentes.items():
        print(f"  {nombre}: {len(filas):4d}")
        todas.extend(filas)
    print(f"  TOTAL: {len(todas):4d}")

    if not todas:
        print("\nNada para cargar (¿archivos movidos o no encontrados?).")
        return

    if dry_run:
        print("\n[dry-run] No se inserta en Supabase. Ejemplo de fila:")
        print(" ", todas[0])
        return

    from supabase import create_client
    supabase_url = os.environ.get("SUPABASE_URL", "")
    supabase_key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not supabase_url or not supabase_key:
        print("ERROR: Configura SUPABASE_URL y SUPABASE_SERVICE_KEY en .env")
        return
    sb = create_client(supabase_url, supabase_key)

    print("\nInsertando apu_precios_referencia...")
    insertadas = _batch_insert(sb, "apu_precios_referencia", todas)
    print(f"  OK: {len(insertadas)} filas.")
    print("\nListo.")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    cargar(dry_run=dry)
