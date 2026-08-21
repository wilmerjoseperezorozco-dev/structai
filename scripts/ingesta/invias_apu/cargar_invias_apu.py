"""
Carga los archivos APU Regionalizados de Referencia de INVIAS a Supabase.

Fuente: https://www.invias.gov.co/publicaciones/4149/analisis-de-precios-unitarios-apu-regionalizados-de-referencia/
Archivos descargados directo de la app real detrás del portal público:
  https://hermes2.invias.gov.co/APUs/Provincias/<anio>_<periodo>/APU_<codigoprovincia>_<DEPTO>__<PROVINCIA>_<anio>_<periodo>.xlsx

Cada archivo Excel de INVIAS trae:
  - 4 hojas catálogo consolidado: MATERIALES, EQUIPO, TRANSPORTE (con código,
    descripción, unidad, precio) y MANO DE OBRA (sin códigos individuales,
    solo valores de referencia SMLMV/jornal — los códigos de mano de obra
    reales, tipo A0030040, solo aparecen dentro de cada hoja de actividad).
  - Cientas de hojas de actividad, una por numeral INVIAS (ej. "610,3"),
    todas con el mismo formato de plantilla FR-APU-1: secciones I.EQUIPO,
    II.MATERIALES, III.TRANSPORTES, IV.MANO DE OBRA, cada una con su
    SUBTOTAL, y una fila TOTAL COSTO DIRECTO al final.

Verificado con datos reales antes de escribir este script (Meta-Ariari vs.
Casanare, numeral 610,3): tanto el precio unitario como el rendimiento de
cada insumo varían por provincia — por eso la "receta" de cada actividad se
guarda en invias_actividad_insumos con provincia_codigo, no es fija a nivel
nacional.

Uso:
  python cargar_invias_apu.py <ruta_a_carpeta_con_xlsx>
  python cargar_invias_apu.py <ruta_a_un_archivo.xlsx>
"""
from __future__ import annotations

import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv
from supabase import Client, create_client

PROJECT_ROOT = Path(__file__).resolve().parents[3]
load_dotenv(PROJECT_ROOT / "apps" / "api" / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = (
    os.environ.get("SUPABASE_SERVICE_KEY")
    or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    or os.environ["SUPABASE_KEY"]
)

# Región natural por código de departamento DANE — dato derivado nuestro, no
# viene de INVIAS. Solo los departamentos que efectivamente vamos a cargar
# necesitan estar aquí; se amplía a medida que se agreguen más regiones.
REGION_POR_DEPTO: dict[str, str] = {
    "50": "Orinoquía",  # Meta
    "81": "Orinoquía",  # Arauca
    "85": "Orinoquía",  # Casanare
    "99": "Orinoquía",  # Vichada
    "94": "Orinoquía",  # Guainía
    "95": "Orinoquía",  # Guaviare
    "08": "Caribe",      # Atlántico
}

FILENAME_RE = re.compile(
    r"^APU_(?P<codigo>\d+)_(?P<depto>.+?)__(?P<provincia>.+?)_(?P<anio>\d{4})_(?P<periodo_n>\d+)\.xlsx$",
    re.IGNORECASE,
)

# Hojas que no son actividades — catálogos, portada, índice, imágenes, etc.
HOJAS_NO_ACTIVIDAD = {
    "PORTADA", "ÍNDICE", "MENÚ", "APU´S", "APU'S", "INSUMO_EQUIPO",
    "INSUMO MATERIALES", "INSUMO_TRANSPORTE", "INSUMO_MANO DE OBRA",
    "IMAGENES_PROVINCIAS", "CLASIFICACIÓN_APU", "HOJA DE CALCULOS",
    "HOJA DE CALCULOS ", "LISTADO DE PROVINCIAS", "MATERIALES", "EQUIPO",
    "MANO DE OBRA", "TRANSPORTE", "APU BASE", "CONSIDERACIONES",
    "CONSIDERACIONES ",
}

# Patrón real de un numeral INVIAS: dígitos separados por comas, ej "610,3",
# "840,11", "201,8,1". Cualquier otra hoja se ignora silenciosamente.
NUMERAL_RE = re.compile(r"^\d+(,\d+)*$")


def _num(v: Any) -> float | None:
    """Convierte a float si es numérico, si no None (celdas de Excel traen
    de todo: fórmulas rotas '#VALUE!', texto, None)."""
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _numeral_coincide(celda_numerica: float, nombre_hoja: str) -> bool:
    """True si una celda que Excel convirtió a número (ej. 730.4) representa
    el mismo numeral que el nombre de hoja con coma (ej. '730,4')."""
    try:
        return abs(celda_numerica - float(nombre_hoja.strip().replace(",", "."))) < 1e-9
    except ValueError:
        return False


def _slug_sin_tildes(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    ).upper()


class CargadorInvias:
    def __init__(self, client: Client):
        self.sb = client
        # Caches en memoria para no reconsultar/re-upsertear el mismo código
        # miles de veces dentro de un mismo archivo.
        self._insumos_vistos: set[str] = set()
        self._actividades_vistas: set[str] = set()

    def cargar_archivo(self, ruta: Path) -> None:
        m = FILENAME_RE.match(ruta.name)
        if not m:
            print(f"  [SALTADO] nombre de archivo no reconocido: {ruta.name}")
            return

        codigo_provincia = m.group("codigo")
        depto = m.group("depto").replace("_", " ").strip()
        provincia = m.group("provincia").replace("_", " ").strip()
        periodo = f"{m.group('anio')}-{int(m.group('periodo_n'))}"
        codigo_depto = codigo_provincia[:2]

        print(f"\n=== {ruta.name} ===")
        print(f"  Provincia {codigo_provincia}: {depto} / {provincia}, periodo {periodo}")

        self._upsert_provincia(codigo_provincia, codigo_depto, depto, provincia)

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            self._cargar_catalogos_insumos(wb, codigo_provincia, periodo)
            n_actividades = self._cargar_actividades(wb, codigo_provincia, periodo)
            print(f"  Actividades cargadas: {n_actividades}")
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # 1. Provincia
    # ------------------------------------------------------------------
    def _upsert_provincia(self, codigo: str, codigo_depto: str, depto: str, provincia: str) -> None:
        region = REGION_POR_DEPTO.get(codigo_depto)
        self.sb.table("invias_provincias").upsert(
            {
                "codigo": codigo,
                "codigo_departamento": codigo_depto,
                "departamento": depto,
                "provincia": provincia,
                "region_natural": region,
            },
            on_conflict="codigo",
        ).execute()

    # ------------------------------------------------------------------
    # 2. Catálogos consolidados: MATERIALES, EQUIPO, TRANSPORTE
    #    (MANO DE OBRA no trae códigos individuales, se ignora aquí — los
    #    códigos de mano de obra se capturan dentro de cada actividad)
    # ------------------------------------------------------------------
    def _cargar_catalogos_insumos(self, wb, codigo_provincia: str, periodo: str) -> None:
        specs = [
            # (nombre_hoja, tipo_insumo, idx_codigo, idx_unidad, idx_insumo, idx_precio, idx_categoria)
            ("MATERIALES", "material", 2, 3, 4, 5, 6),
            ("EQUIPO", "equipo", 2, 3, 4, 6, None),
            ("TRANSPORTE", "transporte", 2, 3, 4, 5, None),
        ]
        for nombre_hoja, tipo_insumo, i_cod, i_und, i_ins, i_precio, i_cat in specs:
            if nombre_hoja not in wb.sheetnames:
                continue
            ws = wb[nombre_hoja]
            insumos_batch: list[dict] = []
            precios_por_codigo: dict[str, dict] = {}
            for row in ws.iter_rows(values_only=True):
                codigo = row[i_cod] if len(row) > i_cod else None
                if not isinstance(codigo, str) or not codigo.strip():
                    continue
                codigo = codigo.strip()
                # Fila de encabezado repetida ("Código") o basura — el código
                # real de INVIAS siempre trae letra + dígitos.
                if not re.match(r"^[A-Za-z]\d", codigo):
                    continue
                descripcion = row[i_ins] if len(row) > i_ins else None
                unidad = row[i_und] if len(row) > i_und else None
                precio = _num(row[i_precio]) if len(row) > i_precio else None
                categoria = row[i_cat] if i_cat is not None and len(row) > i_cat else None
                if not descripcion or precio is None:
                    continue

                if codigo not in self._insumos_vistos:
                    self._insumos_vistos.add(codigo)
                    insumos_batch.append(
                        {
                            "codigo": codigo,
                            "descripcion": str(descripcion).strip(),
                            "unidad": str(unidad).strip() if unidad else None,
                            "tipo_insumo": tipo_insumo,
                            "categoria": str(categoria).strip() if categoria else None,
                        }
                    )
                precios_por_codigo[codigo] = {
                    "insumo_codigo": codigo,
                    "provincia_codigo": codigo_provincia,
                    "periodo": periodo,
                    "precio": precio,
                }

            precios_batch = list(precios_por_codigo.values())
            if insumos_batch:
                self.sb.table("invias_insumos").upsert(insumos_batch, on_conflict="codigo").execute()
            if precios_batch:
                for i in range(0, len(precios_batch), 500):
                    self.sb.table("invias_insumos_precios").upsert(
                        precios_batch[i : i + 500],
                        on_conflict="insumo_codigo,provincia_codigo,periodo",
                    ).execute()
            print(f"  {nombre_hoja}: {len(precios_batch)} precios ({len(insumos_batch)} insumos nuevos)")

    # ------------------------------------------------------------------
    # 3. Hojas de actividad (una por numeral INVIAS)
    # ------------------------------------------------------------------
    def _cargar_actividades(self, wb, codigo_provincia: str, periodo: str) -> int:
        actividades_batch: list[dict] = []
        insumos_linea_batch: list[dict] = []
        costos_batch: list[dict] = []
        insumos_nuevos_mano_obra: list[dict] = []
        saltadas = 0

        for nombre_hoja in wb.sheetnames:
            if nombre_hoja in HOJAS_NO_ACTIVIDAD:
                continue
            if not NUMERAL_RE.match(nombre_hoja.strip()):
                continue

            ws = wb[nombre_hoja]
            rows = list(ws.iter_rows(values_only=True))
            parsed = self._parsear_hoja_actividad(rows, nombre_hoja)
            if parsed is None:
                saltadas += 1
                continue

            numeral, descripcion, unidad, lineas, subtotales, costo_directo = parsed

            if numeral not in self._actividades_vistas:
                self._actividades_vistas.add(numeral)
                actividades_batch.append(
                    {"numeral": numeral, "descripcion": descripcion, "unidad": unidad}
                )

            for linea in lineas:
                insumos_linea_batch.append(
                    {
                        "numeral": numeral,
                        "provincia_codigo": codigo_provincia,
                        "periodo": periodo,
                        "insumo_codigo": linea["codigo"] if linea["codigo"] in self._insumos_vistos else None,
                        "insumo_descripcion": linea["descripcion"],
                        "tipo_insumo": linea["tipo_insumo"],
                        "cantidad_o_rendimiento": linea["cantidad"],
                        "valor_unitario_linea": linea["valor"],
                    }
                )
                # Mano de obra: capturar el insumo en el catálogo maestro
                # también, ya que MANO DE OBRA (consolidada) no trae códigos.
                if linea["tipo_insumo"] == "mano_obra" and linea["codigo"] and linea["codigo"] not in self._insumos_vistos:
                    self._insumos_vistos.add(linea["codigo"])
                    insumos_nuevos_mano_obra.append(
                        {
                            "codigo": linea["codigo"],
                            "descripcion": linea["descripcion"],
                            "unidad": "jornal",
                            "tipo_insumo": "mano_obra",
                            "categoria": None,
                        }
                    )

            if costo_directo is not None:
                costos_batch.append(
                    {
                        "numeral": numeral,
                        "provincia_codigo": codigo_provincia,
                        "periodo": periodo,
                        "costo_equipo": subtotales.get("equipo", 0) or 0,
                        "costo_materiales": subtotales.get("material", 0) or 0,
                        "costo_transporte": subtotales.get("transporte", 0) or 0,
                        "costo_mano_obra": subtotales.get("mano_obra", 0) or 0,
                        "costo_directo_total": costo_directo,
                    }
                )

        if insumos_nuevos_mano_obra:
            self.sb.table("invias_insumos").upsert(
                insumos_nuevos_mano_obra, on_conflict="codigo"
            ).execute()
        if actividades_batch:
            for i in range(0, len(actividades_batch), 300):
                self.sb.table("invias_actividades").upsert(
                    actividades_batch[i : i + 300], on_conflict="numeral"
                ).execute()
        if costos_batch:
            for i in range(0, len(costos_batch), 300):
                self.sb.table("invias_actividad_costos").upsert(
                    costos_batch[i : i + 300],
                    on_conflict="numeral,provincia_codigo,periodo",
                ).execute()
        if insumos_linea_batch:
            for i in range(0, len(insumos_linea_batch), 500):
                self.sb.table("invias_actividad_insumos").insert(
                    insumos_linea_batch[i : i + 500]
                ).execute()

        if saltadas:
            print(f"  [aviso] {saltadas} hojas con nombre de numeral pero formato inesperado, saltadas")
        return len(costos_batch)

    @staticmethod
    def _parsear_hoja_actividad(
        rows: list[tuple], nombre_hoja: str
    ) -> tuple[str, str, str | None, list[dict], dict[str, float], float | None] | None:
        """Parsea una hoja de actividad individual (formato FR-APU-1).

        Estructura confirmada real (verbatim, sección por sección):
          fila ÍTEM: idx1=numeral, idx2=descripción, idx11=unidad
          secciones 'I. EQUIPO' / 'II. MATERIALES' / 'III. TRANSPORTES' /
          'IV. MANO DE OBRA', cada línea: idx1=código, idx2=descripción,
          idx13=Vr.unitario de la línea, idx11 (o idx9 si idx11 vacío) =
          cantidad/rendimiento. Cada sección cierra con fila 'SUBTOTAL $'
          (valor en idx13). Cierra con 'TOTAL COSTO DIRECTO $' (idx13).
        """
        numeral = descripcion = unidad = None
        for row in rows:
            if len(row) <= 2:
                continue
            celda = row[1]
            # Excel a veces auto-convierte un numeral simple ("730,4") a
            # número decimal (730.4) por el separador de coma regional —
            # comparar como string O como número normalizado.
            coincide = (isinstance(celda, str) and celda.strip() == nombre_hoja.strip()) or (
                isinstance(celda, (int, float))
                and _numeral_coincide(celda, nombre_hoja)
            )
            if coincide:
                numeral = nombre_hoja.strip()
                descripcion = str(row[2]).strip() if row[2] else None
                unidad = str(row[11]).strip() if len(row) > 11 and row[11] else None
                break
        if not numeral or not descripcion:
            return None

        secciones = {
            "I. EQUIPO": "equipo",
            "II. MATERIALES": "material",
            "III. TRANSPORTES": "transporte",
            "IV. MANO DE OBRA": "mano_obra",
        }

        lineas: list[dict] = []
        subtotales: dict[str, float] = {}
        costo_directo: float | None = None
        seccion_actual: str | None = None

        for row in rows:
            texto_col1 = row[1] if len(row) > 1 and isinstance(row[1], str) else None

            if texto_col1 in secciones:
                seccion_actual = secciones[texto_col1]
                continue

            if texto_col1 and texto_col1.strip().upper().startswith("SUBTOTAL"):
                if seccion_actual:
                    valor = _num(row[13]) if len(row) > 13 else None
                    if valor is not None:
                        subtotales[seccion_actual] = valor
                seccion_actual = None
                continue

            if texto_col1 and "TOTAL COSTO DIRECTO" in texto_col1.upper():
                costo_directo = _num(row[13]) if len(row) > 13 else None
                continue

            if seccion_actual and texto_col1:
                codigo = texto_col1.strip()
                # línea real de insumo: código empieza con letra + dígito
                # (evita capturar encabezados 'CÓDIGO', notas, etc.)
                if not re.match(r"^[A-Za-z]", codigo) or len(codigo) < 4:
                    continue
                if codigo.upper() in {"CÓDIGO", "CODIGO"}:
                    continue
                desc_linea = row[2] if len(row) > 2 and row[2] else codigo
                valor_linea = _num(row[13]) if len(row) > 13 else None
                cantidad = _num(row[11]) if len(row) > 11 else None
                if cantidad is None:
                    cantidad = _num(row[9]) if len(row) > 9 else None
                if valor_linea is None:
                    continue
                lineas.append(
                    {
                        "codigo": codigo,
                        "descripcion": str(desc_linea).strip(),
                        "tipo_insumo": seccion_actual,
                        "cantidad": cantidad,
                        "valor": valor_linea,
                    }
                )

        return numeral, descripcion, unidad, lineas, subtotales, costo_directo


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    ruta = Path(sys.argv[1])
    archivos = sorted(ruta.glob("*.xlsx")) if ruta.is_dir() else [ruta]
    if not archivos:
        print(f"No se encontraron .xlsx en {ruta}")
        sys.exit(1)

    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    cargador = CargadorInvias(client)
    for archivo in archivos:
        cargador.cargar_archivo(archivo)

    print(f"\nListo. {len(archivos)} archivo(s) procesado(s).")


if __name__ == "__main__":
    main()
