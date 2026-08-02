"""
Generación y parseo de plantillas Excel para cálculo de APU en lote.

Este módulo solo traduce filas de Excel <-> dicts planos — NO calcula
ningún APU. La orquestación (llamar a get_apu()/calcular_apu_dinamico(),
manejar errores por fila, registrar en Supabase) vive en main.py, que ya
tiene esas dependencias cargadas vía el patrón importlib existente.

Dos modos por fila (columna `modo`):
  - "catalogo": actividad_id + cantidad → motor-apu catálogo estático
  - "dinamico": tipo + geometría → motor-apu cálculo por geometría real
"""
from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

MAX_FILAS = 200

COLUMNAS_ENTRADA = [
    "modo", "actividad_id", "cantidad", "tipo", "base_m", "altura_m",
    "longitud_m", "recubrimiento_m", "calidad_concreto", "num_barras",
    "diametro_barra_pulg", "diametro_estribo_pulg", "espaciamiento_estribos_m",
    "gancho_desarrollo_m", "incluye_cara_superior_formaleta", "nota",
]

COLUMNAS_SALIDA = [
    "estado", "descripcion", "unidad", "capitulo", "costo_materiales",
    "costo_mano_obra", "costo_equipo", "costo_directo", "aiu",
    "precio_unitario", "costo_total_fila", "pu_p05", "pu_p95", "pu_std",
    "norma_ref", "uuid_trazabilidad", "error_detalle",
]

_ENCABEZADO_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_ENCABEZADO_FONT = Font(color="FFFFFF", bold=True)
_VERDADEROS = {"true", "verdadero", "si", "sí", "1", "x"}


def _bool_desde_celda(valor: Any) -> bool:
    if isinstance(valor, bool):
        return valor
    if valor is None:
        return False
    return str(valor).strip().lower() in _VERDADEROS


def _num_desde_celda(valor: Any) -> Optional[float]:
    if valor is None or valor == "":
        return None
    return float(valor)


# ── Generar plantilla ─────────────────────────────────────────────────────────

def generar_plantilla(catalogo: list[dict]) -> bytes:
    """catalogo: salida de listar_actividades() — id/descripcion/unidad/capitulo/precio_unitario."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Calculo"
    for col, nombre in enumerate(COLUMNAS_ENTRADA, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.fill = _ENCABEZADO_FILL
        c.font = _ENCABEZADO_FONT
    ws.freeze_panes = "A2"

    ejemplo_catalogo = [
        "catalogo", "C.COL.40X30", 4, "", "", "", "", "", "", "", "", "", "", "", "",
        "Ej: 4 columnas 40x30 del catálogo",
    ]
    ejemplo_dinamico = [
        "dinamico", "", "", "columna", 0.35, 0.35, 3.0, 0.04, "3000", 8, "5", "3", 0.20, 0.10, "FALSE",
        "Ej: columna 35x35, 8 barras #5, estribos #3 c/20",
    ]
    for col, valor in enumerate(ejemplo_catalogo, start=1):
        ws.cell(row=2, column=col, value=valor)
    for col, valor in enumerate(ejemplo_dinamico, start=1):
        ws.cell(row=3, column=col, value=valor)

    dv_modo = DataValidation(type="list", formula1='"catalogo,dinamico"', allow_blank=True)
    dv_modo.add(f"A2:A{MAX_FILAS + 1}")
    ws.add_data_validation(dv_modo)

    dv_tipo = DataValidation(type="list", formula1='"columna,viga"', allow_blank=True)
    dv_tipo.add(f"D2:D{MAX_FILAS + 1}")
    ws.add_data_validation(dv_tipo)

    dv_calidad = DataValidation(type="list", formula1='"2000,2500,3000,4000"', allow_blank=True)
    dv_calidad.add(f"I2:I{MAX_FILAS + 1}")
    ws.add_data_validation(dv_calidad)

    dv_diam = DataValidation(type="list", formula1='"3,4,5,6,7,8"', allow_blank=True)
    dv_diam.add(f"K2:K{MAX_FILAS + 1}")
    dv_diam.add(f"L2:L{MAX_FILAS + 1}")
    ws.add_data_validation(dv_diam)

    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_bool.add(f"O2:O{MAX_FILAS + 1}")
    ws.add_data_validation(dv_bool)

    for col in range(1, len(COLUMNAS_ENTRADA) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16
    ws.column_dimensions["P"].width = 40

    # Hoja de referencia: catálogo válido
    ws2 = wb.create_sheet("Catalogo_APU")
    headers2 = ["actividad_id", "descripcion", "unidad", "capitulo", "precio_unitario"]
    for col, nombre in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=col, value=nombre)
        c.fill = _ENCABEZADO_FILL
        c.font = _ENCABEZADO_FONT
    for row, item in enumerate(catalogo, start=2):
        ws2.cell(row=row, column=1, value=item["id"])
        ws2.cell(row=row, column=2, value=item["descripcion"])
        ws2.cell(row=row, column=3, value=item["unidad"])
        ws2.cell(row=row, column=4, value=item["capitulo"])
        ws2.cell(row=row, column=5, value=item["precio_unitario"])
    for col, ancho in zip("ABCDE", (16, 45, 10, 12, 16)):
        ws2.column_dimensions[col].width = ancho

    # Hoja de instrucciones
    ws3 = wb.create_sheet("Instrucciones")
    instrucciones = [
        "CÓMO LLENAR ESTA PLANTILLA",
        "",
        "1. Use la hoja 'Calculo'. Cada fila es un elemento a calcular.",
        "2. Columna 'modo': 'catalogo' (actividad ya definida) o 'dinamico' (geometría propia).",
        "",
        "MODO 'catalogo':",
        "  - actividad_id: uno de los IDs de la hoja 'Catalogo_APU' (ej: C.COL.40X30).",
        "  - cantidad: número de unidades a calcular.",
        "  - Deje vacías las columnas de geometría (tipo, base_m, altura_m, ...).",
        "",
        "MODO 'dinamico' (columna o viga de concreto reforzado, dimensiones reales):",
        "  - tipo: 'columna' o 'viga'.",
        "  - base_m, altura_m, longitud_m: dimensiones de la sección en metros.",
        "  - recubrimiento_m: opcional, por defecto 0.04 (NSR-10 C.7.7).",
        "  - calidad_concreto: 2000, 2500, 3000 o 4000 PSI (por defecto 3000).",
        "  - num_barras + diametro_barra_pulg: refuerzo longitudinal (opcional).",
        "    diametro_barra_pulg es la clave en octavos de pulgada: 3=3/8\", 4=1/2\", 5=5/8\", 6=3/4\", 7=7/8\", 8=1\".",
        "  - diametro_estribo_pulg + espaciamiento_estribos_m: estribos (opcional).",
        "  - incluye_cara_superior_formaleta: TRUE si la viga es aislada (se forman las 4 caras), si no FALSE.",
        "  - cantidad: número de elementos IGUALES a este (ej: 8 columnas C-1 iguales). Por defecto 1.",
        "",
        "3. Columna 'nota': texto libre suyo para identificar la fila (no se usa para calcular).",
        "4. Máximo 200 filas por archivo.",
        "5. No modifique los nombres de columna de la fila 1.",
        "6. Al subir el archivo calculado en la app, recibirá de vuelta este mismo archivo",
        "   con las columnas de resultado agregadas (costos, precio unitario, trazabilidad NSR-10/NTC).",
        "7. Columna 'estado': OK (calculado), ERROR (fila mal llenada, en rojo — corrija y vuelva a subir),",
        "   OMITIDO (en ámbar — el plan gratis llegó a su límite mensual de cálculos a mitad del archivo;",
        "   active Pro para procesar el resto sin volver a subirlo).",
    ]
    for row, texto in enumerate(instrucciones, start=1):
        c = ws3.cell(row=row, column=1, value=texto)
        if row == 1:
            c.font = Font(bold=True, size=13)
    ws3.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Leer filas subidas ────────────────────────────────────────────────────────

def leer_filas(file_bytes: bytes) -> list[dict]:
    """Lee la hoja 'Calculo' de un .xlsx subido. Devuelve filas en bruto
    (sin validar reglas de negocio — eso lo hace main.py contra el motor real).
    Lanza ValueError con el número de fila si falta 'modo' o el modo es inválido."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Calculo" not in wb.sheetnames:
        raise ValueError("El archivo no tiene una hoja llamada 'Calculo' — use la plantilla descargada, no la edite de estructura.")
    ws = wb["Calculo"]

    header = [c.value for c in ws[1]]
    if header[: len(COLUMNAS_ENTRADA)] != COLUMNAS_ENTRADA:
        raise ValueError("Los encabezados de la hoja 'Calculo' no coinciden con la plantilla — descargue una plantilla nueva.")

    filas: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        valores = [c.value for c in row[: len(COLUMNAS_ENTRADA)]]
        if all(v is None or v == "" for v in valores):
            continue
        if len(filas) >= MAX_FILAS:
            raise ValueError(f"El archivo supera el máximo de {MAX_FILAS} filas.")

        d = dict(zip(COLUMNAS_ENTRADA, valores))
        modo = str(d.get("modo") or "").strip().lower()
        if modo not in ("catalogo", "dinamico"):
            raise ValueError(f"Fila {row_idx}: columna 'modo' debe ser 'catalogo' o 'dinamico', encontrado: '{d.get('modo')}'")

        fila: dict = {
            "_fila_excel": row_idx,
            "modo": modo,
            "nota": d.get("nota") or "",
        }
        if modo == "catalogo":
            fila["actividad_id"] = str(d.get("actividad_id") or "").strip()
            fila["cantidad"] = _num_desde_celda(d.get("cantidad")) or 1.0
        else:
            fila["tipo"] = str(d.get("tipo") or "").strip().lower()
            fila["base_m"] = _num_desde_celda(d.get("base_m"))
            fila["altura_m"] = _num_desde_celda(d.get("altura_m"))
            fila["longitud_m"] = _num_desde_celda(d.get("longitud_m"))
            fila["recubrimiento_m"] = _num_desde_celda(d.get("recubrimiento_m")) or 0.04
            fila["calidad_concreto"] = str(d.get("calidad_concreto") or "3000").strip()
            num_barras = _num_desde_celda(d.get("num_barras"))
            diam_barra = d.get("diametro_barra_pulg")
            fila["refuerzo"] = (
                {"numero_barras": int(num_barras), "diametro_pulg": str(diam_barra).strip()}
                if num_barras and diam_barra else None
            )
            diam_estribo = d.get("diametro_estribo_pulg")
            espac_estribo = _num_desde_celda(d.get("espaciamiento_estribos_m"))
            fila["estribos"] = (
                {
                    "diametro_pulg": str(diam_estribo).strip(),
                    "espaciamiento_m": espac_estribo,
                    "gancho_desarrollo_m": _num_desde_celda(d.get("gancho_desarrollo_m")) or 0.10,
                }
                if diam_estribo and espac_estribo else None
            )
            fila["incluye_cara_superior_formaleta"] = _bool_desde_celda(d.get("incluye_cara_superior_formaleta"))
            fila["cantidad"] = _num_desde_celda(d.get("cantidad")) or 1.0
        filas.append(fila)

    if not filas:
        raise ValueError("El archivo no tiene filas con datos (todas vacías).")
    return filas


# ── Generar workbook de resultado ─────────────────────────────────────────────

def generar_resultado(filas: list[dict], resultados: list[dict]) -> bytes:
    """filas: lo que devolvió leer_filas(). resultados: mismo largo y orden,
    cada dict con claves de COLUMNAS_SALIDA (estado='OK'|'ERROR')."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    columnas = COLUMNAS_ENTRADA + COLUMNAS_SALIDA
    for col, nombre in enumerate(columnas, start=1):
        c = ws.cell(row=1, column=col, value=nombre)
        c.fill = _ENCABEZADO_FILL
        c.font = _ENCABEZADO_FONT
    ws.freeze_panes = "A2"

    fill_error = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    # Amarillo/ámbar, distinto del rojo de error: "OMITIDO" no es un dato mal
    # llenado por el usuario, es una fila que no se calculó porque el plan
    # gratis se quedó sin cupo mensual a mitad del archivo — visualmente debe
    # leerse como "pendiente por plan", no como "corrija esto".
    fill_omitido = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    total = 0.0
    for row_idx, (fila, resultado) in enumerate(zip(filas, resultados), start=2):
        col = 1
        for nombre in COLUMNAS_ENTRADA:
            ws.cell(row=row_idx, column=col, value=fila.get(nombre) if nombre != "modo" else fila.get("modo"))
            col += 1
        for nombre in COLUMNAS_SALIDA:
            ws.cell(row=row_idx, column=col, value=resultado.get(nombre))
            col += 1
        estado = resultado.get("estado")
        if estado == "ERROR":
            for c in range(1, len(columnas) + 1):
                ws.cell(row=row_idx, column=c).fill = fill_error
        elif estado == "OMITIDO":
            for c in range(1, len(columnas) + 1):
                ws.cell(row=row_idx, column=c).fill = fill_omitido
        else:
            total += resultado.get("costo_total_fila") or 0.0

    fila_total = len(filas) + 3
    ws.cell(row=fila_total, column=1, value="TOTAL PRESUPUESTO (filas sin error)").font = Font(bold=True)
    idx_costo_total = len(COLUMNAS_ENTRADA) + COLUMNAS_SALIDA.index("costo_total_fila") + 1
    ws.cell(row=fila_total, column=idx_costo_total, value=round(total, 2)).font = Font(bold=True)

    for col in range(1, len(columnas) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16
    ws.column_dimensions[get_column_letter_of("nota", COLUMNAS_ENTRADA)].width = 35
    ws.column_dimensions[get_column_letter_of("descripcion", COLUMNAS_ENTRADA + COLUMNAS_SALIDA)].width = 40
    ws.column_dimensions[get_column_letter_of("uuid_trazabilidad", COLUMNAS_ENTRADA + COLUMNAS_SALIDA)].width = 38
    ws.column_dimensions[get_column_letter_of("error_detalle", COLUMNAS_ENTRADA + COLUMNAS_SALIDA)].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_column_letter_of(nombre_columna: str, columnas: list[str]) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(columnas.index(nombre_columna) + 1)
